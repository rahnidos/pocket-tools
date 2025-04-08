from openpyxl import load_workbook
import sys 
from datetime import datetime, timedelta
import requests



def get_rate(date):
     api_url = f"https://api.nbp.pl/api/exchangerates/rates/A/EUR/{date}/?format=json"
     response = requests.get(api_url)
     if response.status_code == 200:
         data = response.json()
         rate = data['rates'][0]['mid']
         return rate
     if response.status_code == 404:
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        previous_day = date_obj - timedelta(days=1)
        #print(previous_day.strftime('%Y-%m-%d'))
        return get_rate(previous_day.strftime('%Y-%m-%d'))
     else:
        print(f"Error: {response.status_code}")
        return None
     
try:                                                                                                                             
        file_path=sys.argv[1]                                                                                                       
except IndexError:                                                                                                               
        print ("no path provided")
        sys.exit(1)   


workbook = load_workbook(filename=file_path)
sheet = workbook.active
last_date= None



for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        value  = cell.value
        if isinstance(value,str):
            date_obj = datetime.strptime(value, '%d %b %Y')
        elif isinstance(value, datetime):
            date_obj = value
        else:
            print("Invalid date format")
            continue
        previous_day = date_obj - timedelta(days=1)
        formatted_date = previous_day.strftime('%Y-%m-%d')
        if last_date != formatted_date:
            rate = get_rate(formatted_date)
            if rate is not None:
                cell.offset(column=7 - cell.column).value = rate
                last_date = formatted_date
            else:
                print(f"Rate not found for date: {formatted_date}")
        else:
            cell.offset(column=7 - cell.column).value = rate
            last_date = formatted_date
        print(f"Date: {formatted_date}, Rate: {rate}")
workbook.save(file_path)